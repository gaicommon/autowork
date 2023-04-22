from openpyxl import load_workbook
import os.path
from openpyxl.styles import NamedStyle, Font, colors
from ProjVar.var import *
from Util.fromattime import *


class Excel:
    def __init__(self, excel_file_path):
        self.sheet = None
        self.column_num = None
        self.row_num = None
        if os.path.exists(excel_file_path):
            self.excel_file_path = excel_file_path
            self.wb = load_workbook(self.excel_file_path)
        else:
            print("%s e文件的路劲不存在，请重新设定！" % excel_file_path)

    # 通过sheet名字来获取操作的sheet
    def set_sheet_by_name(self, sheet_name):
        if sheet_name in self.wb.sheetnames:
            self.sheet = self.wb[sheet_name]
        else:
            print("%s sheet不存在，请重新指定！" % sheet_name)

    # 通过序号来获取操作的sheet
    def set_sheet_by_index(self, index):
        if isinstance(index, int) and 1 <= index <= len(self.get_all_sheet_names()):
            sheet_name = self.get_all_sheet_names()[index - 1]
            self.sheet = self.wb[sheet_name]
        else:
            print("%s sheet 序号不存在，请重新设定" % index)

    # 获取当前sheet和title的名称
    def get_current_sheet_name(self):
        return self.sheet.title

    # 获取所有sheet的名称
    def get_all_sheet_names(self):
        return self.wb.sheetnames

    # 获取sheet的总行数，从0开始，返回list
    def get_rows_object(self):
        return list(self.sheet.rows)

    # 获取sheet的总列数，从0开始，返回list
    def get_cols_object(self):
        return list(self.sheet.columns)

    # 获取某行的对象，第一行从0开始
    def get_row(self, row_no):
        return self.get_rows_object()[row_no]

    def get_excel_value_list(self, row_no):
        row = self.get_row(row_no)
        row_value_dic = {}
        if self.column_num:
            for i in range(self.column_num):
                row_value_dic[row[i].value]=i
        return row_value_dic

    # 获取某一列对象，第一列从0开始
    def get_col(self, col_no):
        return self.get_cols_object()[col_no]

    # 获取某个单元格对象
    def get_cell_value(self, row_no, col_no):
        if isinstance(row_no, int) and isinstance(col_no, int) and \
                1 <= row_no <= len(self.get_rows_object()) and \
                1 <= row_no <= len(self.get_cols_object()):
            return self.sheet.cell(row=row_no, column=col_no).value
        else:
            print("%s,%s 行号或者列号不存在，请重新设定行号或者列表读取！" % (row_no, col_no))

    # 给某一个单元格写入指定内容，行号、列号从1开始
    # 调用此方法时，excel不要处于打开状态
    def write_cell_value(self, row_no, col_no, value, color=None):
        if isinstance(row_no, int) and isinstance(col_no, int):
            if color is None:
                font = Font(bold=False, size=10, color=colors.BLACK)
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).value = value
            elif color == "green":
                font = Font(bold=True, size=13, color=colors.GREEN)
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).value = value
            elif color == "red":
                font = Font(bold=True, size=13, color=colors.RED)
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).value = value
            self.wb.save(self.excel_file_path)
        else:
            print("%s,%s 行号或者列号不是数字，请重新设定行号或者列表读取！" % (row_no, col_no))

    def write_current_time(self, row_no, col_no):
        if isinstance(row_no, int) and isinstance(col_no, int):
            self.sheet.cell(row=row_no, column=col_no).value = get_current_date_and_time()
            self.wb.save(self.excel_file_path)


if __name__ == "__main__":
    excel_file_path = ProjDirPath + r"\TestData\126邮箱联系人.xlsx"
    # print(excel_file_path )
    excel_obj = Excel(excel_file_path)
    # Excel("e:\\a.xlsx")    #测试路劲不存在的情况
    # excel_obj.set_sheet_by_name("测试邮件")
    # excel_obj.set_sheet_by_name("测试邮件1111")
    excel_obj.set_sheet_by_index(1)
    # print(excel_obj.get_current_sheet_name())
    # excel_obj.set_sheet_by_index(5)
    # print(excel_obj.get_rows_object())
    # print(excel_obj.get_cols_object())
    # print(excel_obj.get_row(2))
    print(excel_obj.get_cell_value(2, 2))
    print(excel_obj.write_cell_value(5, 7, "hello~~"))
    print(excel_obj.write_current_time(6, 8))
